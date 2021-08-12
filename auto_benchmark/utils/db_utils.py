import sys
# print(sys.path)
# from auto_benchmark.utils import DataInsert
import logging
import pymysql
import xlwt
import xlrd
import openpyxl
from fileio import load
import pandas as pd
from .db import Connection, Info
from xlutils.copy import copy


def get_data_from_db(path):
    conn = Connection(Info.DBInfo(), Info.UserInfo())
    cursor = conn.connection().cursor()
    tabel = conn.dbinfo.datatable
    sql = conn.dbinfo.sql
    cursor.execute(sql)
    results = cursor.fetchall()
    fields = cursor.description
    workbook = xlwt.Workbook(encoding='utf-8')
    sheet = workbook.add_sheet(path.split('.')[0].split("/")[1],
                               cell_overwrite_ok=True)
    for field in range(0, len(fields)):
        sheet.write(0, field, fields[field][0])
    for row in range(1, len(results) + 1):
        for col in range(0, len(fields)):
            sheet.write(row, col, results[row - 1][col])
    workbook.save(path)


def build_sh_to_xls(sh_file, db_file, target_file):

    workbook = xlwt.Workbook(encoding="utf-8")  # 创建 xls 文件,可被复写

    sheet = workbook.add_sheet("from_shfile", cell_overwrite_ok=True)

    fields = [
        '模型序号', 'DataSource', 'FrameName', 'ModelName', 'BENCHMARK_FS',
        'NumCards', 'speed', 'TagOrBranch', 'ExecDate', 'Node'
    ]

    for field in range(0, len(fields)):
        sheet.write(0, field, fields[field])

    data = xlrd.open_workbook(db_file)
    table = data.sheets()[0]

    sh_file = open(sh_file, 'r')
    text_lines = sh_file.readlines()
    gsh_row = 0
    for line in text_lines:
        if not line.startswith("#"):
            gsh_row = gsh_row + 1
            line_ = line.split(" ")
            for item in line_:
                if 'DETAIL_INFO' in item:
                    detail_info = item.split("*")
                    framework = detail_info[0].split("=")[1]
                    gpus = detail_info[1]
                    model = detail_info[2]
                elif "BENCHMARK_FS" in item:
                    benchmark_fs = item.split("=")[1]
                else:
                    pass
            sheet.write(gsh_row, 0, gsh_row)
            sheet.write(gsh_row, 2, framework)
            sheet.write(gsh_row, 3, model)
            sheet.write(gsh_row, 4, benchmark_fs)
            sheet.write(gsh_row, 5, gpus)

            target_frame = framework
            target_NumCards = gpus
            target_Modelname = model + "*" + benchmark_fs

            for row in range(1, table.nrows):
                # frame，modelname(包含测试方式)，卡数
                if target_frame == table.row_values(row)[2] and \
                 target_Modelname == table.row_values(row)[3] and  \
                        target_NumCards == str(int(table.row_values(row)[8])):
                    sheet.write(gsh_row, 1, table.row_values(row)[1])
                    sheet.write(gsh_row, 6, table.row_values(row)[10])  # speed
                    sheet.write(gsh_row, 7,
                                table.row_values(row)[18])  # TagBranch
                    sheet.write(gsh_row, 8,
                                table.row_values(row)[22])  # ExecDate
                    sheet.write(gsh_row, 8, table.row_values(row)[9])  # Node
    workbook.save(target_file)


def build_fail_sh(framelist=None,
                  datasource=None,
                  shsource=None,
                  shtarget=None):
    data = xlrd.open_workbook(datasource)
    table = data.sheets()[0]
    sh_file = open(shsource, 'r')
    text_lines = sh_file.readlines()
    for line in text_lines:
        flag = 0
        if not line.startswith("#"):
            line_ = line.split(" ")
            for item in line_:
                if 'BENCHMARK_FS' in item:
                    benchmark_fs = item.split('=')[-1]
                if 'DETAIL_INFO' in item:
                    model_name = item.split('*')[-1]
            model_info = model_name + "*" + benchmark_fs
            for row in range(0, table.nrows):
                if table.row_values(row)[3] != model_info:
                    flag = 1
                else:
                    flag = 0
                    break
            if flag:
                with open(shtarget, "a+") as f:
                    f.writelines(line)

# 从拉下来的数据表格生成根据分级排序的模型
# 特点1:按照模型优先级
# 特点2:dummy和rel一起同一框架在一起，以框架粒度。
def write_to_xls(framelist, source_file, save_path):
    workbook = xlwt.Workbook(encoding="utf-8")  # 创建 xls 文件,可被复写
    sheet = workbook.add_sheet(save_path.split('.')[0].split('/')[1],
                               cell_overwrite_ok=True)
    fields = ['模型序号', '框架', '测试方式', '模型', '卡数', '版本', 's/iter']
    for field in range(0, len(fields)):
        sheet.write(0, field, fields[field])
    data = xlrd.open_workbook(source_file)
    table = data.sheets()[0]

    frames = load(framelist, file_format='yaml')
    grow_frame = 0
    for model in frames:
        grow_model = 0
        for row in range(0, table.nrows):
            if table.row_values(row)[2] == model:
                grow_frame += 1
                grow_model += 1
                for col in range(len(table.row_values(row))):
                    if col == 2:
                        sheet.write(grow_frame, 0, grow_model)
                        sheet.write(grow_frame, 1, table.row_values(row)[col])
                    elif col == 3:
                        fs = table.row_values(row)[col].split("*")
                        sheet.write(grow_frame, 3, fs[0])
                        sheet.write(grow_frame, 2, fs[1])
                    elif col == 8:
                        sheet.write(grow_frame, 4, table.row_values(row)[col])
                    elif col == 10:
                        sheet.write(grow_frame, 6, table.row_values(row)[col])
                    elif col == 18:
                        sheet.write(grow_frame, 5, table.row_values(row)[col])
            else:
                print(f"skip model {model}")
                continue
    workbook.save(save_path)

#每个模型dummy和real放到一起，以模型粒度
def to_target_format(framelist, source, target):
    frame_set = []
    source_target = pd.read_excel(source)
    frames = load(framelist, file_format='yaml')
    for frame in frames:
        temp_frame = source_target[(source_target['框架']==frame)].sort_values(['模型'])
        # for index in range(len(temp_frame)):
        #     temp_frame.iloc[index, 0] = index + 1
        frame_set.append(temp_frame)
    target_table = pd.concat(frame_set)
    target_table.to_excel(target)


def from_dh_sort_table(framelist, source_file, save_path):
    frames = load(framelist, file_format='yaml')
    # read xls
    data = xlrd.open_workbook(source_file)
    table = data.sheets()[0]

    # build xls for save
    workbook = xlwt.Workbook(encoding="utf-8")  # 创建 xls 文件,可被复写
    sheet = workbook.add_sheet(save_path.split('.')[0].split('/')[1],
                               cell_overwrite_ok=True)

    for col in range(0, table.ncols):
        sheet.write(0, col, table.row_values(0)[col])
    grow = 0
    for frame in frames:
        for row in range(1, table.nrows):
            # print(table.row_values(row))
            if table.row_values(row)[0] == frame:
                grow = grow + 1
                for col in range(0, table.ncols):
                    sheet.write(grow, col, table.row_values(row)[col])
    workbook.save(save_path)


def add_new_data_to_old_excel(old_file, new_file, save_path):
    # read xls
    old_data = xlrd.open_workbook(old_file)
    old_table = old_data.sheets()[0]

    new_data=copy(old_data)
    new_table = new_data.get_sheet(0)

    temp_data = xlrd.open_workbook(new_file)
    temp_table = temp_data.sheets()[0]

    #build xls for save
    workbook = xlwt.Workbook(encoding="utf-8")  # 创建 xls 文件,可被复写
    sheet = workbook.add_sheet(save_path.split('.')[0].split('/')[1], cell_overwrite_ok=True)

    for col in range(0, old_table.ncols):
        sheet.write(0, col, old_table.row_values(0)[col])

    for row in range(1, old_table.nrows):
        for row_temp in range(1, temp_table.nrows):
            if old_table.row_values(row)[1] == \
                temp_table.row_values(row_temp)[2] and  old_table.row_values(row)[4] == temp_table.row_values(row_temp)[4]:
                if "REAL_ITERTIME" in temp_table.row_values(row_temp)[3]:
                    new_table.write(row, old_table.ncols, temp_table.row_values(row_temp)[7])
                elif "DUMMY_ITERTIME" in temp_table.row_values(row_temp)[3]:
                    new_table.write(row, old_table.ncols+1, temp_table.row_values(row_temp)[7])
    new_data.save(save_path)


if __name__ == "__main__":
    # get_data_from_db()
    # build_fail_sh()
    # write_to_xls()
    # to_target_format()
    sh_file = "/mnt/lustre/wuwenli1/workspace/benchmark_platform/auto-bench/auto_benchmark/wzh/mmpose_wzh.sh"
    sh_file = "/mnt/lustre/wuwenli1/workspace/benchmark_platform/auto-bench/auto_benchmark/wzh/mmpose_dd.sh"
    sh_file = "/mnt/lustre/wuwenli1/workspace/benchmark_platform/auto-bench/auto_benchmark/wzh/merged_sketch.sh"
    db_file = "/mnt/lustre/wuwenli1/workspace/benchmark_platform/auto-bench/auto_benchmark/data_from_db_sketch1.4.xls"
    target_file = "/mnt/lustre/wuwenli1/workspace/benchmark_platform/auto-bench/auto_benchmark/dbfile_shfile_.xls"
    sh_file = "/mnt/lustre/wuwenli1/workspace/benchmark_platform/auto-bench/auto_benchmark/merged_3.sh"
    build_sh_to_xls(sh_file, db_file, target_file)
